"""
@Author: haoran.xu
@Modified: Python Automation Assistant
"""
import os
import csv
import pathlib
from typing import Tuple, Optional

from PIL import Image, ImageChops, ImageDraw, ImageFont  # 引入 ImageDraw 和 ImageFont
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# --- 配置常量 ---
BROWSER_PATH = r"C:\PythonWorkSpace\DiffAutoMator\tools"
os.environ['PLAYWRIGHT_BROWSERS_PATH'] = BROWSER_PATH

IMG_TEMP = 'temp_raw.png'
IMG_FINAL = 'temp_optimized.png'
EXCEL_OUT = 'test_excel.xlsx'
GLOBAL_FONT = "ＭＳ Ｐゴシック"

# --- 限制常量 ---
# Excel/OpenPyXL 对图片高度的安全限制 (像素)
MAX_HEIGHT_LIMIT = 25000
filtered_filenames = []


# --- 样式管理 ---
class Styles:
    BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    FONT_HEADER = Font(name=GLOBAL_FONT, color="FFFFFF", bold=True)
    FONT_NORMAL = Font(name=GLOBAL_FONT, color="000000", bold=False)

    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')


# --- 核心功能函数 ---

def init_workbook() -> Tuple[Workbook, Worksheet]:
    """初始化Excel工作簿及表头"""
    wb = Workbook()
    ws = wb.active
    ws.title = "diff一览"

    headers = ["No.", "ファイル名", "フォルダ", "比較結果", "備考", "問題有無"]
    cols_width = [4, 50, 35, 50, 60, 20]

    for col_idx, (title, width) in enumerate(zip(headers, cols_width), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = Styles.FILL_HEADER
        cell.font = Styles.FONT_HEADER
        cell.border = Styles.BORDER_THIN
        cell.alignment = Styles.ALIGN_CENTER
        ws.column_dimensions[chr(64 + col_idx)].width = width

    return wb, ws


def process_text_data(txt_path: str, ws: Worksheet) -> int:
    """读取TXT数据并写入Excel，返回最后一行行号"""
    if not os.path.exists(txt_path):
        print(f"[Err] 文件不存在: {txt_path}")
        return 1

    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            rows = list(reader)

        if not rows:
            return 1

        current_row = 2
        for idx, data in enumerate(rows[1:], 1):
            fname = data[0] if len(data) > 0 else ""
            folder = data[1] if len(data) > 1 else ""
            result = data[2] if len(data) > 2 else ""

            is_target = '$' in fname
            remark = "対象外" if is_target else ""

            row_vals = [idx, fname, folder, result, remark, ""]
            aligns = ['right', 'left', 'left', 'left', 'center', 'center']

            for c_idx, (val, align_type) in enumerate(zip(row_vals, aligns), 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.font = Styles.FONT_NORMAL
                cell.border = Styles.BORDER_THIN
                cell.alignment = Styles.ALIGN_CENTER if align_type == 'center' else Styles.ALIGN_LEFT

                if is_target:
                    cell.fill = Styles.FILL_GRAY

            if not is_target and fname:
                filtered_filenames.append(fname)

            current_row += 1

        print(f"[Info] 文本数据处理完成，共 {len(rows) - 1} 行")
        print(len(filtered_filenames))
        print(filtered_filenames)
        return current_row - 1

    except Exception as e:
        print(f"[Err] 数据处理失败: {e}")
        return 1


# --- 字体和水印相关函数 ---

def get_system_font(size: int = 80):
    """尝试获取与 GLOBAL_FONT (MS P Gothic) 匹配的系统字体"""

    # 字体文件列表 (MS P Gothic 常见的几个文件名)
    font_files = ['msgothic.ttc']

    # 常见的 Windows 字体路径
    win_fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')

    for font_file in font_files:
        path = os.path.join(win_fonts_dir, font_file)
        if os.path.exists(path):
            try:
                # 加载字体并设置大小
                return ImageFont.truetype(path, size)
            except IOError:
                # 字体文件存在但无法加载
                continue

    # 所有尝试失败
    return None


def add_watermark(img_path: str, text: str = "画像が大きすぎます"):
    """
    给图片添加现代优雅的红色半透明水印。
    特点：无红框，大号，居中，高透明度。
    """
    try:
        img = Image.open(img_path).convert("RGBA")

        # 创建绘制对象，与原图大小相同，用于叠加水印
        overlay = Image.new("RGBA", img.size, (255, 255, 255, 0))
        draw = ImageDraw.Draw(overlay)

        w, h = img.size

        # 1. 尝试获取系统字体
        font = get_system_font(size=80)

        # 2. 决定显示内容和颜色 (使用高透明度软红)
        if font:
            watermark_text = text  # 中文
            text_fill = (255, 0, 0, 100)  # 纯红，透明度100 (最大255)
        else:
            watermark_text = "IMAGE TOO LARGE - CROPPED"  # 英文回退
            text_fill = (255, 0, 0, 100)
            font = ImageFont.load_default()  # 使用默认字体

        # 3. 计算文字居中位置
        try:
            # textbbox 替代 textsize，更准确地计算边界
            bbox = draw.textbbox((0, 0), watermark_text, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
        except AttributeError:
            # Fallback for older PIL versions
            text_w, text_h = draw.textsize(watermark_text, font=font)

        text_x = (w - text_w) // 2
        text_y = (h - text_h) // 2

        # 4. 绘制文字 (无红框)
        draw.text((text_x, text_y), watermark_text, fill=text_fill, font=font)

        # 5. 合并图片
        img = Image.alpha_composite(img, overlay)
        img = img.convert("RGB")  # 转回 RGB 兼容 Excel
        img.save(img_path)
        print(f"[Info] 已添加水印: {text} ")

    except Exception as e:
        print(f"[Err] 水印添加失败: {e}")
        if os.path.exists(img_path):
            print("[Warn] 保持原始图片，未添加水印。")


# --- 截图和优化函数 ---

def capture_screenshot(html_path: str, output_path: str) -> bool:
    """
    截图逻辑：
    1. 模拟 1920x1080 窗口
    2. 缩放 80%
    3. 检查高度：如果超过阈值 (25000px)，仅截取一页并打水印；否则长截图。
    """
    abs_path = pathlib.Path(html_path).resolve()
    if not abs_path.exists():
        print("[Err] HTML文件未找到")
        return False

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(viewport={'width': 1920, 'height': 1080})
            page.goto(f"file://{abs_path}")
            page.wait_for_load_state('networkidle')

            # --- 修改点1：设置页面缩放为 80% ---
            # 这比后期图片 resize 更清晰，且能截取到更宽的内容
            page.evaluate("document.body.style.zoom = '80%'")

            # 检查页面高度
            scroll_height = page.evaluate("document.body.scrollHeight")
            print(f"[Info] 页面高度: {scroll_height}px (限制: {MAX_HEIGHT_LIMIT}px)")

            is_oversized = scroll_height > MAX_HEIGHT_LIMIT

            if is_oversized:
                # 图片过大：只截取首屏 (full_page=False)
                print("[Warn] 图片高度超过Excel限制，执行单页截图...")
                page.screenshot(path=output_path, full_page=False)
            else:
                # 正常范围：长截图
                page.screenshot(path=output_path, full_page=True)
            browser.close()

        # 如果是超大图片，截完图后追加水印
        if is_oversized:
            add_watermark(output_path)

        return True
    except Exception as e:
        print(f"[Err] 截图失败: {e}")
        return False


def optimize_image(input_path: str, output_path: str, scale: float = 1.0) -> Optional[str]:
    """
    智能去白边并（可选）缩放图片。
    注意：由于已经在浏览器进行了80%缩放，这里的默认 scale 建议设为 1.0，除非需要再次缩小。
    """
    if not os.path.exists(input_path):
        return None

    try:
        img = Image.open(input_path)

        # 1. 计算内容边界（去白边）
        bg = Image.new(img.mode, img.size, (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()

        if bbox:
            # 四周保留细微空白 (Padding)
            margin = 10

            # 计算带 margin 的新坐标，并防止超出原图边界
            left = bbox[0]
            top = max(0, bbox[1] - margin)
            right = min(img.width, bbox[2] + margin)
            bottom = min(img.height, bbox[3] + margin)

            img = img.crop((left, top, right, bottom))

        # 2. 缩放 (默认为 1.0，因为已在浏览器端缩放 80%)
        if scale != 1.0:
            new_size = (int(img.width * scale), int(img.height * scale))
            img = img.resize(new_size, Image.Resampling.LANCZOS)

        img.save(output_path)
        return output_path
    except Exception as e:
        print(f"[Err] 图片优化异常: {e}")
        return None


def cleanup_temp_files():
    for f in [IMG_TEMP, IMG_FINAL]:
        if os.path.exists(f):
            try:
                os.remove(f)
            except:
                pass


# --- 主程序入口 ---

def main():
    print("--- 自动化报告生成开始 ---")

    # 1. 获取输入
    txt_path = input("TXT路径: ").strip('"\'')
    html_path = input("HTML路径: ").strip('"\'')

    # 2. Excel 数据处理
    wb, ws = init_workbook()
    last_row = process_text_data(txt_path, ws)

    # 3. 截图与图像处理
    print("[Info] 正在生成截图...")
    # capture_screenshot 中已完成 80% 缩放和尺寸判断
    if capture_screenshot(html_path, IMG_TEMP):
        # scale 保持 1.0，因为浏览器端已缩放
        final_img_path = optimize_image(IMG_TEMP, IMG_FINAL, scale=0.9)

        # 4. 插入图片
        if final_img_path:
            insert_pos = f'B{last_row + 2}'
            ws.add_image(ExcelImage(final_img_path), insert_pos)
            print(f"[Info] 图片已插入位置: {insert_pos}")

    # 5. 保存与清理
    try:
        wb.save(EXCEL_OUT)
        print(f"[Success] 文件已生成: {EXCEL_OUT}")
    except Exception as e:
        print(f"[Err] 保存Excel失败: {e}")
    finally:
        cleanup_temp_files()


if __name__ == "__main__":
    main()