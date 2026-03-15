"""
@Author: haoran.xu
"""
import os
import csv
import pathlib
import math
from datetime import datetime
from typing import Tuple, Optional, List

from PIL import Image, ImageChops, ImageDraw, ImageFont
from playwright.sync_api import sync_playwright

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from constants import EXCEL_DIR, PNG_CACHE_DIR

import sys
from pathlib import Path

def resource_path(relative_path):
    """获取资源文件的绝对路径，用于PyInstaller打包后的程序"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = Path(__file__).parent
    return str(Path(base_path) / relative_path)

BASE_DIR = os.getcwd()
BROWSER_PATH = resource_path("tools")
os.environ['PLAYWRIGHT_BROWSERS_PATH'] = BROWSER_PATH

os.makedirs(PNG_CACHE_DIR, exist_ok=True)
IMG_TEMP = os.path.join(PNG_CACHE_DIR, 'temp_raw.png')
IMG_FINAL = os.path.join(PNG_CACHE_DIR, 'temp_optimized.png')
os.makedirs(EXCEL_DIR, exist_ok=True)
EXCEL_OUT = os.path.join(EXCEL_DIR, f"diff_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
GLOBAL_FONT = "ＭＳ Ｐゴシック"

MAX_HEIGHT_LIMIT = 25000

filtered_filenames = []
FILTERED_EXCLUSIONS = {'logstash-logback-encoder'}


class Styles:
    BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    FONT_HEADER = Font(name=GLOBAL_FONT, color="FFFFFF", bold=True)
    FONT_NORMAL = Font(name=GLOBAL_FONT, color="000000", bold=False)

    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

    FONT_FILE_TITLE_SIMPLE = Font(name=GLOBAL_FONT, color="000000", bold=True)


def init_workbook() -> Tuple[Workbook, Worksheet]:
    """初始化Excel工作簿及表头"""
    wb = Workbook()
    ws = wb.active
    ws.title = "diff一览"

    headers = ["No.", "ファイル名", "フォルダ", "比較結果", "備考", "問題有無"]
    cols_width = [4, 50, 35, 50, 60, 20]

    for col_idx, (title, width) in enumerate(zip(headers, cols_width), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = Styles.FILL_HEADER
        cell.font = Styles.FONT_HEADER
        cell.border = Styles.BORDER_THIN
        cell.alignment = Styles.ALIGN_CENTER
        ws.column_dimensions[chr(64 + col_idx)].width = width

    return wb, ws


def process_text_data(txt_path: str, ws: Worksheet) -> int:
    """读取TXT数据并写入Excel，返回最后一行行号"""
    global filtered_filenames
    filtered_filenames = []

    if not os.path.exists(txt_path):
        print(f"[Err] 文件不存在: {txt_path}")
        return 1

    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            rows = list(reader)

        if not rows:
            return 1

        current_row = 2
        for idx, data in enumerate(rows[1:], 1):
            fname = data[0] if len(data) > 0 else ""
            folder = data[1] if len(data) > 1 else ""
            result = data[2] if len(data) > 2 else ""

            is_target = '$' in fname
            remark = "対象外" if is_target else ""

            row_vals = [idx, fname, folder, result, remark, ""]
            aligns = ['right', 'left', 'left', 'left', 'center', 'center']

            for c_idx, (val, align_type) in enumerate(zip(row_vals, aligns), 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.font = Styles.FONT_NORMAL
                cell.border = Styles.BORDER_THIN
                cell.alignment = Styles.ALIGN_LEFT

                if is_target:
                    cell.fill = Styles.FILL_GRAY

            # 逻辑：非target、有文件名、且不包含过滤关键字，加入待处理列表
            if not is_target and fname and not any(exclusion in fname for exclusion in FILTERED_EXCLUSIONS):
                filtered_filenames.append(fname)

            current_row += 1

        print(f"[Info] 文本数据处理完成，共 {len(rows) - 1} 行")
        print(f"[Info] 待提取详情的文件数: {len(filtered_filenames)}")
        return current_row - 1

    except Exception as e:
        print(f"[Err] 数据处理失败: {e}")
        return 1


# --- 截图和优化函数 ---

def capture_screenshot(html_path: str, output_path: str, page_evaluate: int = 100) -> bool:
    """
    截图逻辑：
    1. 模拟 1920x1080 窗口
    2. 缩放 80%
    3. 检查高度并截图
    """
    abs_path = pathlib.Path(html_path).resolve()
    if not abs_path.exists():
        print(f"[Err] HTML文件未找到: {abs_path}")
        return False

    try:
        with sync_playwright() as p:
            # 尝试启动浏览器，如果失败则尝试安装后再启动
            try:
                browser = p.chromium.launch(headless=True)
            except Exception as launch_error:
                print(f"[Info] 浏览器启动失败，尝试安装浏览器... {launch_error}")
                try:
                    # 尝试安装浏览器
                    p.chromium.download()
                    browser = p.chromium.launch(headless=True)
                except Exception as install_error:
                    print(f"[Err] 无法下载或启动浏览器: {install_error}")
                    return False

            page = browser.new_page(viewport={'width': 1920, 'height': 1080})
            page.goto(f"file://{abs_path}")
            page.wait_for_load_state('networkidle')

            # 页面缩放为 80%
            page.evaluate(f"document.body.style.zoom = '{page_evaluate}%'")

            # 检查页面高度
            scroll_height = page.evaluate("document.body.scrollHeight")
            is_oversized = scroll_height > MAX_HEIGHT_LIMIT

            if is_oversized:
                print(f"[Warn] 图片高度 {scroll_height}px 超过Excel限制，仅截取首屏...")
                page.screenshot(path=output_path, full_page=False)
            else:
                page.screenshot(path=output_path, full_page=True)
            browser.close()

        return True
    except Exception as e:
        print(f"[Err] 截图失败: {e}")
        return False


def optimize_image(input_path: str, output_path: str, scale: float = 1.0) -> Optional[str]:
    """智能去白边并（可选）缩放图片"""
    if not os.path.exists(input_path):
        return None

    try:
        img = Image.open(input_path)
        white_color = 255 if img.mode == 'L' else (255, 255, 255)
        if img.mode == 'RGBA':
            white_color = (255, 255, 255, 255)

        # 1. 去白边
        bg = Image.new(img.mode, img.size, white_color)
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()

        if bbox:
            # 预留白边
            margin = 0
            left = bbox[0]
            top = max(0, bbox[1] - margin)
            right = min(img.width, bbox[2] + margin)
            bottom = min(img.height, bbox[3] + margin)
            img = img.crop((left, top, right, bottom))

        # 2. 缩放
        if scale != 1.0:
            new_size = (int(img.width * scale), int(img.height * scale))
            img = img.resize(new_size, Image.Resampling.LANCZOS)

        img.save(output_path)
        return output_path
    except Exception as e:
        print(f"[Err] 图片优化异常: {e}")
        return None


def calculate_rows_needed(img_path: str) -> int:
    """根据图片高度计算Excel需要跳过的行数"""
    try:
        with Image.open(img_path) as img:
            height = img.height
            # Excel 默认行高约 15点 ≈ 20像素。
            # 为了保险起见，按 18像素/行 计算，并额外加 2 行缓冲
            rows = math.ceil(height / 18)
            return rows + 1
    except Exception:
        return 18  # 如果读取失败，返回默认值


def cleanup_temp_files(files_to_remove: List[str]):
    """清理指定的文件列表"""
    for f in files_to_remove:
        if os.path.exists(f):
            try:
                os.remove(f)
            except Exception as e:
                print(f"[Warn] 删除文件失败 {f}: {e}")


def cleanup_png_cache():
    """清理整个PNG缓存目录中的所有PNG文件"""
    if os.path.exists(PNG_CACHE_DIR):
        for file in os.listdir(PNG_CACHE_DIR):
            file_path = os.path.join(PNG_CACHE_DIR, file)
            if file_path.endswith('.png') and os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    print(f"[Info] Cleaned cache file: {file_path}")
                except Exception as e:
                    print(f"[Warn] Failed to delete cache file {file_path}: {e}")


def find_latest_comparison_dir():
    """查找merge_files目录下最新的comparison子文件夹"""
    # 使用constants中的TARGET_DIR
    from constants import TARGET_DIR
    merge_files_dir = os.path.join(os.getcwd(), str(TARGET_DIR))
    if not os.path.exists(merge_files_dir):
        print(f"[Err] 未找到 merge_files 目录: {merge_files_dir}")
        return None

    comparison_dirs = []
    try:
        for item in os.listdir(merge_files_dir):
            item_path = os.path.join(merge_files_dir, item)
            if os.path.isdir(item_path) and item.startswith("comparison_"):
                comparison_dirs.append(item_path)
    except Exception as e:
        print(f"[Err] 遍历目录失败: {e}")
        return None

    if not comparison_dirs:
        print(f"[Err] 未找到 comparison_ 开头的子文件夹")
        return None

    # 按修改时间排序，获取最新的文件夹
    latest_dir = max(comparison_dirs, key=os.path.getmtime)
    print(f"[Info] 找到最新的比较目录: {latest_dir}")
    return latest_dir


def find_files_in_dir(directory):
    """在指定目录中查找txt和html文件"""
    txt_files = []
    html_files = []

    for file in os.listdir(directory):
        file_path = os.path.join(directory, file)
        if file.endswith('.txt') and file != 'file_list_to_extract.txt':
            txt_files.append(file_path)
        elif file.endswith('.html') and 'vs' in file.lower():
            html_files.append(file_path)

    return txt_files, html_files


def extract_file_search_step():
    """抽取的文件查找和验证步骤"""
    print("--- 自动化报告生成开始 ---")

    # 1. 自动查找文件
    latest_dir = find_latest_comparison_dir()
    if not latest_dir:
        return None, None, None

    txt_files, html_files = find_files_in_dir(latest_dir)

    if not txt_files:
        print(f"[Err] 未找到符合条件的TXT文件")
        return None, None, None
    if not html_files:
        print(f"[Err] 未找到包含 'vs' 的HTML文件")
        return None, None, None

    # 默认取第一个找到的文件
    txt_path = txt_files[0]
    html_path = html_files[0]

    print(f"[Info] 使用TXT文件: {os.path.basename(txt_path)}")
    print(f"[Info] 使用主HTML文件: {os.path.basename(html_path)}")

    return latest_dir, txt_path, html_path


def extract_data_processing_step(txt_path):
    """抽取的Excel初始化和数据处理步骤"""
    # 2. Excel 数据处理
    wb, ws = init_workbook()
    last_row = process_text_data(txt_path, ws)
    return wb, ws, last_row


def extract_main_screenshot_step(html_path, ws, last_row):
    """抽取的主截图处理步骤"""
    # 3. 处理主截图 (VS)
    print("[Info] 正在处理主对比图...")
    current_row = last_row + 5  # 默认值
    if capture_screenshot(html_path, IMG_TEMP, 90):
        final_img_path = optimize_image(IMG_TEMP, IMG_FINAL, scale=1.0)
        if final_img_path:
            insert_pos = f'B{last_row + 2}'
            ws.add_image(ExcelImage(final_img_path), insert_pos)

            # 计算图片占用的行数，更新 current_row
            rows_needed = calculate_rows_needed(final_img_path)
            current_row = last_row + 2 + rows_needed
            print(f"[Info] 主图片插入于 {insert_pos}, 占用约 {rows_needed} 行")
        else:
            current_row = last_row + 5
    else:
        current_row = last_row + 5

    return current_row


def extract_detail_screenshot_step(latest_dir, ws, current_row_start):
    """抽取的详情截图处理步骤"""
    # 4. 循环处理 filtered_filenames 对应的详情截图
    print("--- 开始处理详情文件 ---")

    # 使用 copy() 进行遍历，以便安全修改原列表
    filenames_to_process = filtered_filenames.copy()
    temp_files_to_clean = [IMG_TEMP, IMG_FINAL]
    current_row = current_row_start
    # is_first_screenshot = True

    for filename in filenames_to_process:
        # 构建预期的HTML文件路径
        expected_html_name = f"{filename}_merge_report.html"
        expected_html_path = os.path.join(latest_dir, expected_html_name)

        if os.path.exists(expected_html_path):
            print(f"[Info] 匹配到文件: {filename} -> {expected_html_name}")

            # 4.1 写入文件名标题
            # 空一行 (current_row 已指向空行)
            title_cell = ws.cell(row=current_row, column=2, value=filename)
            title_cell.font = Styles.FONT_FILE_TITLE_SIMPLE
            title_cell.alignment = Styles.ALIGN_LEFT
            current_row += 1

            # 4.2 生成唯一的临时文件名
            timestamp = datetime.now().strftime('%H%M%S_%f')
            temp_raw = os.path.join(PNG_CACHE_DIR, f"temp_{filename}_{timestamp}.png")
            temp_opt = os.path.join(PNG_CACHE_DIR, f"opt_{filename}_{timestamp}.png")
            temp_files_to_clean.extend([temp_raw, temp_opt])

            # if is_first_screenshot:
            # 第一张截图，传入缩放 90
            #     success = capture_screenshot(expected_html_path, temp_raw, 90)
            #     is_first_screenshot = False
            # else:
            #     success = capture_screenshot(expected_html_path, temp_raw)

            # 4.3 截图与插入
            if capture_screenshot(expected_html_path, temp_raw):
                final_sub_path = optimize_image(temp_raw, temp_opt, scale=1.0)

                if final_sub_path:
                    insert_pos = f'B{current_row}'

                    # 用excel的缩放，减少画质影响
                    # img = ExcelImage(final_sub_path)
                    # img.width = img.width * 0.9
                    # img.height = img.height * 0.9
                    ws.add_image(ExcelImage(final_sub_path), insert_pos)

                    # 动态计算下一行的位置，避免图片重叠
                    rows_occupied = calculate_rows_needed(final_sub_path)
                    current_row += rows_occupied

                    print(f"       已插入图片，预留 {rows_occupied} 行")
                else:
                    current_row += 2
            else:
                ws.cell(row=current_row, column=2, value="[截图失败]").font = Styles.FONT_NORMAL
                current_row += 2

            # 4.4 从全局列表中移除已处理的项
            if filename in filtered_filenames:
                filtered_filenames.remove(filename)

            # 插入分隔空行
            current_row += 0

        else:
            # 如果没找到HTML，不删除列表中的值，仅跳过
            pass

    return current_row, temp_files_to_clean


def find_jar_file_paths():
    """
    查找./ExtractedDiffs目录下最新txt文件中与filtered_filenames集合匹配的JAR文件绝对路径
    返回jar_files集合，其中键为filtered_filenames中的原始名称，值为对应的绝对路径
    """
    import os
    from pathlib import Path

    jar_files = {}  # 使用字典存储对应关系：filtered_filenames中的名称 -> 绝对路径
    extracted_diffs_dir = Path("./ExtractedDiffs")

    if not extracted_diffs_dir.exists():
        print(f"[Info] 未找到 ExtractedDiffs 目录: {extracted_diffs_dir}")
        return jar_files

    # 查找ExtractedDiffs目录下最新的txt文件
    txt_files = []
    for root, dirs, files in os.walk(extracted_diffs_dir):
        for file in files:
            if file.endswith('.txt') and 'Extracted_File_Paths' in file:
                txt_files.append(Path(root) / file)

    if not txt_files:
        print(f"[Info] 未找到包含 'Extracted_File_Paths' 的txt文件")
        return jar_files

    # 找到最新的txt文件
    latest_txt_file = max(txt_files, key=os.path.getctime)
    print(f"[Info] 找到最新的提取路径文件: {latest_txt_file}")

    # 读取txt文件内容
    try:
        with open(latest_txt_file, 'r', encoding='utf-8') as f:
            content = f.readlines()
    except Exception as e:
        print(f"[Err] 读取文件失败: {e}")
        return jar_files

    # 从filtered_filenames中筛选出JAR文件
    jar_filenames = [name for name in filtered_filenames if name.endswith('.jar')]
    print(f"[Info] filtered_filenames中的JAR文件: {jar_filenames}")

    # 在txt文件内容中查找匹配的路径
    for jar_name in jar_filenames:
        for line in content:
            line = line.strip()
            # 检查该行是否以jar_name结尾
            if line.endswith(jar_name):
                jar_files[jar_name] = line  # 存储对应关系：原始名称 -> 绝对路径
                print(f"[Info] 找到匹配的JAR文件路径: {jar_name} -> {line}")
                break  # 找到匹配后跳出内层循环，继续下一个jar_name

    return jar_files


def create_worksheet_content(latest_dir, ws, txt_path, html_path):
    """创建工作表内容的公共方法"""
    from datetime import datetime

    # 1. Excel 数据处理
    last_row = process_text_data(txt_path, ws)

    # 2. 处理主截图 (VS)
    print("[Info] 正在处理主对比图...")
    current_row = last_row + 5  # 默认值
    if capture_screenshot(html_path, IMG_TEMP, 90):
        final_img_path = optimize_image(IMG_TEMP, IMG_FINAL, scale=1.0)
        if final_img_path:
            insert_pos = f'B{last_row + 2}'
            ws.add_image(ExcelImage(final_img_path), insert_pos)

            # 计算图片占用的行数，更新 current_row
            rows_needed = calculate_rows_needed(final_img_path)
            current_row = last_row + 2 + rows_needed
            print(f"[Info] 主图片插入于 {insert_pos}, 占用约 {rows_needed} 行")
        else:
            current_row = last_row + 5
    else:
        current_row = last_row + 5

    # 3. 循环处理 filtered_filenames 对应的详情截图
    print("--- 开始处理详情文件 ---")

    # 使用 copy() 进行遍历，以便安全修改原列表
    filenames_to_process = filtered_filenames.copy()
    temp_files_to_clean = [IMG_TEMP, IMG_FINAL]

    for filename in filenames_to_process:
        # 构建预期的HTML文件路径
        expected_html_name = f"{filename}_merge_report.html"
        expected_html_path = os.path.join(latest_dir, expected_html_name)

        if os.path.exists(expected_html_path):
            print(f"[Info] 匹配到文件: {filename} -> {expected_html_name}")

            # 3.1 写入文件名标题
            # 空一行 (current_row 已指向空行)
            title_cell = ws.cell(row=current_row, column=2, value=filename)
            title_cell.font = Styles.FONT_FILE_TITLE_SIMPLE
            title_cell.alignment = Styles.ALIGN_LEFT
            current_row += 1

            # 3.2 生成唯一的临时文件名
            timestamp = datetime.now().strftime('%H%M%S_%f')
            temp_raw = os.path.join(PNG_CACHE_DIR, f"temp_{filename}_{timestamp}.png")
            temp_opt = os.path.join(PNG_CACHE_DIR, f"opt_{filename}_{timestamp}.png")
            temp_files_to_clean.extend([temp_raw, temp_opt])

            # 3.3 截图与插入
            if capture_screenshot(expected_html_path, temp_raw):
                final_sub_path = optimize_image(temp_raw, temp_opt, scale=1.0)

                if final_sub_path:
                    insert_pos = f'B{current_row}'
                    ws.add_image(ExcelImage(final_sub_path), insert_pos)

                    # 动态计算下一行的位置，避免图片重叠
                    rows_occupied = calculate_rows_needed(final_sub_path)
                    current_row += rows_occupied

                    print(f"       已插入图片，预留 {rows_occupied} 行")
                else:
                    current_row += 2
            else:
                ws.cell(row=current_row, column=2, value="[截图失败]").font = Styles.FONT_NORMAL
                current_row += 2

            # 3.4 从全局列表中移除已处理的项
            if filename in filtered_filenames:
                filtered_filenames.remove(filename)

            # 插入分隔空行
            current_row += 0

        else:
            # 如果没找到HTML，不删除列表中的值，仅跳过
            pass

    return current_row, temp_files_to_clean


def add_sheet_to_existing_excel(sheet_name_suffix="", sheet_name=""):
    """在现有Excel文件中添加新的工作表，而不是覆盖"""
    from openpyxl import load_workbook
    import os
    from datetime import datetime

    print(f"--- 添加新工作表到现有Excel文件 (工作表: {sheet_name_suffix}) ---")

    # 1. 自动查找文件
    latest_dir = find_latest_comparison_dir()
    if not latest_dir:
        return

    txt_files, html_files = find_files_in_dir(latest_dir)

    if not txt_files:
        print(f"[Err] 未找到符合条件的TXT文件")
        return
    if not html_files:
        print(f"[Err] 未找到包含 'vs' 的HTML文件")
        return

    # 默认取第一个找到的文件
    txt_path = txt_files[0]
    html_path = html_files[0]

    print(f"[Info] 使用TXT文件: {os.path.basename(txt_path)}")
    print(f"[Info] 使用主HTML文件: {os.path.basename(html_path)}")

    # 2. 查找最新的Excel文件
    excel_files = [f for f in os.listdir(EXCEL_DIR) if f.startswith("diff_report_") and f.endswith(".xlsx")]
    if not excel_files:
        print(f"[Err] 未找到现有的Excel报告文件，无法添加新工作表")
        return

    # 按修改时间排序，获取最新的Excel文件
    excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(EXCEL_DIR, x)), reverse=True)
    excel_file_path = os.path.join(EXCEL_DIR, excel_files[0])

    if not os.path.exists(excel_file_path):
        print(f"[Err] Excel文件不存在: {excel_file_path}")
        return

    try:
        wb = load_workbook(excel_file_path)
        # 创建新工作表
        new_sheet_name = f"{sheet_name}" if sheet_name_suffix else f"jar_diff_{datetime.now().strftime('%H%M%S')}"
        # 确保工作表名称不超过31个字符且不包含非法字符
        new_sheet_name = new_sheet_name[:31].replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        ws = wb.create_sheet(title=new_sheet_name)

        headers = ["No.", "ファイル名", "フォルダ", "比較結果", "備考", "問題有無"]
        cols_width = [4, 50, 35, 50, 60, 20]

        for col_idx, (title, width) in enumerate(zip(headers, cols_width), 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = Styles.FILL_HEADER
            cell.font = Styles.FONT_HEADER
            cell.border = Styles.BORDER_THIN
            cell.alignment = Styles.ALIGN_CENTER
            ws.column_dimensions[chr(64 + col_idx)].width = width

        print(f"[Info] 在Excel文件中添加新工作表: {excel_file_path}, 工作表名: {new_sheet_name}")
    except Exception as e:
        print(f"[Err] 加载Excel文件失败: {e}")
        return

    # 3. 创建工作表内容
    current_row, temp_files_to_clean = create_worksheet_content(latest_dir, ws, txt_path, html_path)

    # 4. 保存与清理
    try:
        wb.save(excel_file_path)
        print(f"--- 新工作表添加完成 ---")
        print(f"[Success] Excel已更新: {excel_file_path}, 工作表: {new_sheet_name}")
        print(f"[Info] 剩余未处理文件: {filtered_filenames}")
    except Exception as e:
        print(f"[Err] 保存Excel失败: {e}")
    finally:
        # 清理临时文件列表中的文件
        cleanup_temp_files(temp_files_to_clean)
        # 清理整个PNG缓存目录
        cleanup_png_cache()


def extract_save_and_cleanup_step(wb, temp_files_to_clean):
    """抽取的保存和清理步骤"""
    # 5. 保存与清理
    try:
        wb.save(EXCEL_OUT)
        print(f"--- 处理结束 ---")
        print(f"[Success] Excel已生成: {EXCEL_OUT}")
        print(f"[Info] 剩余未处理文件: {filtered_filenames}")
    except Exception as e:
        print(f"[Err] 保存Excel失败: {e}")
    finally:
        # 清理临时文件列表中的文件
        cleanup_temp_files(temp_files_to_clean)
        # 清理整个PNG缓存目录
        cleanup_png_cache()


# --- 主程序入口 ---

def main():
    # 步骤1: 文件查找和验证
    latest_dir, txt_path, html_path = extract_file_search_step()
    if latest_dir is None:
        return

    # 步骤2: Excel初始化和数据处理
    wb, ws, last_row = extract_data_processing_step(txt_path)

    # 步骤3: 处理主截图
    current_row = extract_main_screenshot_step(html_path, ws, last_row)

    # 步骤4: 处理详情截图
    current_row, temp_files_to_clean = extract_detail_screenshot_step(latest_dir, ws, current_row)

    # 步骤5: 保存和清理
    extract_save_and_cleanup_step(wb, temp_files_to_clean)


if __name__ == "__main__":
    main()